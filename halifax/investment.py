import math
import openpyxl
from re import findall
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.webdriver import WebDriver

from utils import delay


def halifax_investment(driver: WebDriver, output_xlsx):
    url = 'https://www.investments.halifax.co.uk/investment-trust-centre/'
    WAIT = WebDriverWait(driver, 10)
    driver.get(url)

    total = driver.find_element(By.XPATH, '//*[@id="itsTotal"]')
    total = total.text.replace(',', '')
    total_funds = int(findall(r'[0-9]+', total)[0])
    total_pages = math.ceil(total_funds / 100)

    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb['Investment']
    sheet_iter = 2

    print(f'[halifax] Total IT = {total_funds}')

    for page in range(1, total_pages + 1):
        data = []
        current_page = f'https://www.investments.halifax.co.uk/modules/investment-trusts/result/?limit=100&offset={page}&orderField=UnitNameLong&orderType=asc'
        driver.get(current_page)
        print(f'[#] IT Page {page}/{total_pages} [#]')
        WAIT.until(EC.presence_of_element_located(
            (By.XPATH, '/html/body/div[1]/table/tbody/tr')))
        tr = driver.find_elements(By.XPATH, '/html/body/div[1]/table/tbody/tr')
        even_tr = tr[::2]

        for row in even_tr:
            name = row.find_element(By.XPATH, './td[1]/p/a').text.strip()
            url = row.find_element(
                By.XPATH, './td[1]/p/a').get_attribute('href')

            isin = ""
            if url:
                isin_match = findall(r"[A-Z0-9]{12}", url)
                if len(isin_match) > 0:
                    isin = isin_match[0]

            data.append({'name': name, 'url': url, 'isin': isin})

        for output in data:
            ws.cell(sheet_iter, 1).value = output['name']
            ws.cell(sheet_iter, 2).value = output["isin"]
            fund_url = output["url"]
            c = ws.cell(sheet_iter, 3, fund_url)
            c.hyperlink = fund_url
            c.style = "Hyperlink"

            sheet_iter += 1
        wb.save(output_xlsx)
        delay(0.5, 2)
    wb.save(output_xlsx)
    wb.close()
